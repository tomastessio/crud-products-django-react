from decimal import Decimal, InvalidOperation
from django.db import transaction
from django.http import HttpResponse
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from .models import Article
from .serializers import ArticleSerializer
import openpyxl
from openpyxl import Workbook


class ArticleViewSet(viewsets.ModelViewSet):

    queryset = Article.objects.all().order_by("id")
    serializer_class = ArticleSerializer
    
    @action(detail=False, methods=["post"], url_path="import")
    def import_excel(self, request):
        
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "Falta el archivo (campo 'file')."}, status=400)
        try:
            wb = openpyxl.load_workbook(file, data_only=True)
        except Exception as e:
            return Response({"detail": f"Archivo inválido: {e}"}, status=400)
        
        ws = wb.active
        headers = {}
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        
        for idx, name in enumerate(header_row):
            if not name: continue
            key = str(name).strip().lower()
            if key in ("code", "codigo"): headers["code"] = idx
            elif key in ("description", "descripcion", "descripción"): headers["description"] = idx
            elif key in ("price", "precio"): headers["price"] = idx
        
        required = {"code", "description", "price"}
        
        if not required.issubset(headers.keys()):
            return Response({"detail": "Encabezados requeridos no encontrados. Usa: code/codigo, description/descripcion, price/precio"}, status=400)
        
        created, updated, errors = 0, 0, []
        
        with transaction.atomic():
            for r, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row: continue
                try:
                    code = str(row[headers["code"]]).strip()
                    description = str(row[headers["description"]]).strip()
                    raw_price = row[headers["price"]]
                    
                    if raw_price is None or code == "" or description == "":
                        continue
                    
                    if isinstance(raw_price, str): raw_price = raw_price.replace(",", ".")
                    
                    price = Decimal(str(raw_price))
                    
                    obj, created_flag = Article.objects.update_or_create(
                        code=code, defaults={"description": description, "price": price},
                    )
                    
                    created += int(created_flag); updated += int(not created_flag)
                except (InvalidOperation, ValueError) as e:
                    errors.append({"row": r, "error": f"Precio inválido: {e}"})
                
                except Exception as e:
                    errors.append({"row": r, "error": str(e)})
        
        return Response({"created": created, "updated": updated, "errors": errors})
    
    
    @action(detail=False, methods=["get"], url_path="export")
    def export_excel(self, request):
        
        wb = Workbook(); ws = wb.active; ws.title = "Articles"; ws.append(["code","description","price"])
        
        for a in Article.objects.order_by("id").all(): ws.append([a.code, a.description, float(a.price)])
        
        from io import BytesIO
        
        stream = BytesIO(); wb.save(stream); stream.seek(0)
        
        resp = HttpResponse(stream.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
        resp["Content-Disposition"] = 'attachment; filename="articles.xlsx"'
        
        return resp
